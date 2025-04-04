def WorkOfExcel(Column_find,fruit,Update_price):
    import openpyxl
    book = openpyxl.load_workbook("C:\\Users\\Ajith\\Downloads\\download.xlsx")
    sheet = book.active

    a = {}

    for i in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=i).value == Column_find:
            a["column_value"] = i

    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row=i, column=j).value == fruit:
                a["row_value"] = i

    sheet.cell(row=a["row_value"], column=a["column_value"]).value = Update_price
    book.save("C:\\Users\\Ajith\\Downloads\\download4.xlsx")


from selenium import webdriver
from selenium.webdriver.common.by import By
driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.maximize_window()


fruit = "Apple"
Column_find = "price"
Update_price = "700"

#1. download a file
driver.find_element(By.CSS_SELECTOR, "#downloadButton").click()

#2. edit Excel file on top of the program
WorkOfExcel(Column_find,fruit,Update_price)

#3. upload a file
driver.find_element(By.XPATH, "//input[@type='file']").send_keys("C:\\Users\\Ajith\\Downloads\\download4.xlsx")

#validation
# check Excel value and webpage values are same
column_value = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
actual = driver.find_element(By.XPATH, "//div[text()='"+fruit+"']/parent ::div/parent ::div/div[@id='cell-"+column_value+"-undefined']/div").text
print(actual)
assert actual == Update_price

''' Excel file didn't save same location'''