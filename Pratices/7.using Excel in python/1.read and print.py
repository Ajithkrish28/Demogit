'''
handling excel with python

import openpyxl ( excel access toll in this module )
'''

import openpyxl

#open and load the excell we use load work book these are in openpyxl module
book = openpyxl.load_workbook("C:\\Users\\Ajith\\Desktop\\pratics.xlsx")

# excell have many working page, so we activate that page , so, we use active module and store in one variable
sheet = book.active

# print the values
# excell are pointed in rows and columns they are called cell
# so, mention that cell in rows and columns
cell = sheet.cell(row=1, column=2)   # now we store that cell in cell variable

print(cell.value)       # print cell variable values, we use value (it's use for get the value from cell )

'''we get value so, use value'''
# update the values
sheet.cell(row=1, column=1).value = "Ajith"  # we mention which cell we update

print(sheet.cell(row=1,column=1).value)

print(sheet.max_row)  # print a maximum row count

print(sheet.max_column)  # print a maximum column count