import openpyxl

book = openpyxl.load_workbook("C:\\Users\\Ajith\\Desktop\\pratcis.xlsx")
sheet = book.active
a={}
'''
i take one by one values in rows 
and column 1 and rows values == "testcase2"
that row in all column values are get
'''
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "testcase2":
        for j in range(2, sheet.max_column+1):
            a[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value   # rows=1 because all keys are in first row

print(a)
