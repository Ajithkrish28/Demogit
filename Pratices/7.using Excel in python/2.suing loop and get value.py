import openpyxl

book = openpyxl.load_workbook("C:\\Users\\Ajith\\Desktop\\pratcis.xlsx")
sheet = book.active

# it's first column in all row values are printed
for i in range(1, sheet.max_row+1):
    print(sheet.cell(row=i, column=1).value)


print("-------")

# i loop is take's one by one cell  and j loop all are executed and came to i
# j loop is take one by one, once completed ang to i loop

for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value)

print("----------")

# we use condition
# i loop is take's one by one cell  and j loop all are executed and came to i
# and which row's values are == testcase2 then nxt next step execute
# j loop is take one by one, once completed ang to i loop
for i in range(1, sheet.max_row+1):
  if sheet.cell(row=i, column=1).value == "testcase2":
      for j in range(2, sheet.max_column+1):
          print(sheet.cell(row=i, column=j).value)